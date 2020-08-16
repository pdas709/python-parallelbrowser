import openpyxl
from random import randint



class ReadWriteExcel:

  def readandADDCertificate(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 1)
    return cell_obj.value

  def editCertification(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 2)
    return cell_obj.value

  def university(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 3)
    return cell_obj.value

  def degree(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 4)
    return cell_obj.value

  def company(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 5)
    return cell_obj.value  

  def role(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 6)
    return cell_obj.value 

  def job_description(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 7)
    return cell_obj.value 

  def profile_pictureSection(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 8)
    return cell_obj.value 

  def summary_field(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 9)
    return cell_obj.value   

  def min_expected_rate(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 10)
    return cell_obj.value

  def city(self):
    path = "C:\\Users\\decontractor\\Downloads\\TestCases_Candidate_Profile_RGP\\candidate_profile\\testingdata\\Book432.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    
    cell_obj = sheet_obj.cell(row = randint(2,13), column = 11)
    return cell_obj.value   

   